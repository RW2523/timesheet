"""
Report service — Phase 8.
Generates 8-sheet Excel workbook using openpyxl.
"""
import os
import logging
from datetime import datetime
from typing import Optional

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.db.models import (
    GeneratedReport, BatchUpload, UploadedFile, TimesheetSubmission,
    TimesheetEntry, ValidationError, Employee, Vendor, PayrollResult,
    gen_uuid,
)
from app.services.storage_service import StorageService

logger = logging.getLogger(__name__)

# Colour palette
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
GREEN_FILL = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
BLUE_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.storage = StorageService()

    def generate_batch_report(self, batch_id: str) -> GeneratedReport:
        batch = self.db.query(BatchUpload).filter(BatchUpload.id == batch_id).first()
        if not batch:
            raise ValueError(f"Batch {batch_id} not found")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._sheet_batch_summary(wb, batch)
        self._sheet_file_inventory(wb, batch_id)
        self._sheet_extracted_entries(wb, batch_id)
        self._sheet_validation_exceptions(wb, batch_id)
        self._sheet_employee_summary(wb, batch_id)
        self._sheet_vendor_summary(wb, batch_id)
        self._sheet_payroll_report(wb, batch_id)
        self._sheet_adp_export(wb, batch_id)

        # Save file
        reports_dir = self.storage.reports_dir(batch_id)
        file_name = f"timesheets_report_{batch.source_name.replace('.zip', '')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(reports_dir, file_name)
        wb.save(file_path)

        report = GeneratedReport(
            id=gen_uuid(),
            batch_id=batch_id,
            report_type="BATCH_FULL",
            file_name=file_name,
            file_path=file_path,
            created_at=datetime.utcnow(),
        )
        self.db.add(report)
        self.db.commit()
        logger.info(f"[{batch_id}] Report saved: {file_path}")
        return report

    # ── Sheet 1: Batch Summary ─────────────────────────────────────────────────

    def _sheet_batch_summary(self, wb: Workbook, batch: BatchUpload) -> None:
        ws = wb.create_sheet("01_Batch_Summary")
        rows = [
            ("Batch ID", batch.id),
            ("Source File", batch.source_name),
            ("Status", batch.status),
            ("Total Files", batch.total_files),
            ("Processed", batch.processed_files),
            ("Failed", batch.failed_files),
            ("Ignored/Noise", batch.ignored_files),
            ("Duplicates", batch.duplicate_files),
            ("Needs Review", batch.review_required_files),
            ("Payroll Ready", batch.payroll_ready_count),
            ("Created At", str(batch.created_at)),
        ]
        for r, (label, value) in enumerate(rows, 1):
            ws.cell(r, 1, label).font = BOLD
            ws.cell(r, 2, str(value) if value is not None else "")
        self._auto_width(ws)

    # ── Sheet 2: File Inventory ────────────────────────────────────────────────

    def _sheet_file_inventory(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("02_File_Inventory")
        headers = [
            "Folder", "File Name", "Extension", "Size (KB)",
            "Detected Employee", "Matched Employee ID", "Match Status",
            "OCR Required", "Duplicate", "Noise", "Status",
        ]
        self._write_header_row(ws, headers)

        files = self.db.query(UploadedFile).filter(UploadedFile.batch_id == batch_id).all()
        for r, f in enumerate(files, 2):
            ws.cell(r, 1, f.folder_path or "")
            ws.cell(r, 2, f.file_name)
            ws.cell(r, 3, f.file_ext or "")
            ws.cell(r, 4, round((f.file_size_bytes or 0) / 1024, 1))
            ws.cell(r, 5, f.detected_employee_name or "")
            ws.cell(r, 6, f.matched_employee_id or "")
            ws.cell(r, 7, f.match_status)
            ws.cell(r, 8, "Yes" if f.ocr_required else "No")
            ws.cell(r, 9, "Yes" if f.is_duplicate else "No")
            ws.cell(r, 10, "Yes" if f.is_noise_file else "No")
            ws.cell(r, 11, f.processing_status)

            if f.is_duplicate or f.processing_status == "FAILED":
                self._fill_row(ws, r, len(headers), RED_FILL)
            elif f.processing_status == "NEEDS_REVIEW":
                self._fill_row(ws, r, len(headers), YELLOW_FILL)

        self._auto_width(ws)

    # ── Sheet 3: Extracted Entries ─────────────────────────────────────────────

    def _sheet_extracted_entries(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("03_Extracted_Entries")
        headers = [
            "Employee ID", "Work Date", "Day", "In Time", "Out Time",
            "Break (min)", "Entered Hours", "Calculated Hours",
            "Regular Hours", "Overtime Hours", "Entry Type", "Leave Type",
            "Holiday", "Validation Status",
        ]
        self._write_header_row(ws, headers)

        entries = (
            self.db.query(TimesheetEntry)
            .join(TimesheetSubmission, TimesheetEntry.submission_id == TimesheetSubmission.id)
            .filter(TimesheetSubmission.batch_id == batch_id)
            .order_by(TimesheetEntry.employee_id, TimesheetEntry.work_date)
            .all()
        )

        for r, e in enumerate(entries, 2):
            ws.cell(r, 1, e.employee_id or "")
            ws.cell(r, 2, str(e.work_date) if e.work_date else "")
            ws.cell(r, 3, e.day_of_week or "")
            ws.cell(r, 4, str(e.in_time) if e.in_time else "")
            ws.cell(r, 5, str(e.out_time) if e.out_time else "")
            ws.cell(r, 6, e.break_minutes or 0)
            ws.cell(r, 7, float(e.entered_hours) if e.entered_hours else "")
            ws.cell(r, 8, float(e.calculated_hours) if e.calculated_hours else "")
            ws.cell(r, 9, float(e.regular_hours) if e.regular_hours else 0)
            ws.cell(r, 10, float(e.overtime_hours) if e.overtime_hours else 0)
            ws.cell(r, 11, e.entry_type or "")
            ws.cell(r, 12, e.leave_type or "")
            ws.cell(r, 13, "Yes" if e.is_holiday else "No")
            ws.cell(r, 14, e.validation_status)

            if e.validation_status == "FAILED":
                self._fill_row(ws, r, len(headers), RED_FILL)
            elif e.is_holiday:
                self._fill_row(ws, r, len(headers), YELLOW_FILL)

        self._auto_width(ws)

    # ── Sheet 4: Validation Exceptions ────────────────────────────────────────

    def _sheet_validation_exceptions(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("04_Validation_Exceptions")
        headers = [
            "Severity", "Rule Code", "Employee ID", "File",
            "Message", "Expected", "Actual", "Action Required", "Status",
        ]
        self._write_header_row(ws, headers)

        errors = (
            self.db.query(ValidationError)
            .filter(ValidationError.batch_id == batch_id)
            .order_by(ValidationError.severity, ValidationError.created_at)
            .all()
        )

        sev_fill = {"BLOCKER": RED_FILL, "ERROR": RED_FILL, "WARNING": YELLOW_FILL, "INFO": GREEN_FILL}

        for r, e in enumerate(errors, 2):
            ws.cell(r, 1, e.severity)
            ws.cell(r, 2, e.rule_code)
            ws.cell(r, 3, str(e.employee_id) if e.employee_id else "")
            ws.cell(r, 4, str(e.file_id) if e.file_id else "")
            ws.cell(r, 5, e.message)
            ws.cell(r, 6, e.expected_value or "")
            ws.cell(r, 7, e.actual_value or "")
            ws.cell(r, 8, e.action_required or "")
            ws.cell(r, 9, e.status)
            fill = sev_fill.get(e.severity, GREEN_FILL)
            self._fill_row(ws, r, len(headers), fill)

        self._auto_width(ws)

    # ── Sheet 5: Employee Summary ──────────────────────────────────────────────

    def _sheet_employee_summary(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("05_Employee_Summary")
        headers = [
            "Employee ID", "Employee Name", "Total Entries", "Total Regular Hours",
            "Total Overtime Hours", "Leave Days", "Approval Status",
            "Validation Status", "Payroll Status",
        ]
        self._write_header_row(ws, headers)

        submissions = (
            self.db.query(TimesheetSubmission)
            .filter(TimesheetSubmission.batch_id == batch_id)
            .all()
        )

        for r, sub in enumerate(submissions, 2):
            emp = self.db.query(Employee).filter(Employee.id == sub.employee_id).first()
            entries = self.db.query(TimesheetEntry).filter(TimesheetEntry.submission_id == sub.id).all()
            total_reg = sum(float(e.regular_hours or 0) for e in entries)
            total_ot = sum(float(e.overtime_hours or 0) for e in entries)
            leave_days = sum(1 for e in entries if e.entry_type == "LEAVE")

            ws.cell(r, 1, sub.employee_id or "")
            ws.cell(r, 2, emp.full_name if emp else "Unknown")
            ws.cell(r, 3, len(entries))
            ws.cell(r, 4, round(total_reg, 2))
            ws.cell(r, 5, round(total_ot, 2))
            ws.cell(r, 6, leave_days)
            ws.cell(r, 7, sub.approval_status)
            ws.cell(r, 8, sub.validation_status)
            ws.cell(r, 9, sub.payroll_status)

            if sub.payroll_status == "NOT_READY":
                self._fill_row(ws, r, len(headers), RED_FILL)
            elif sub.payroll_status == "READY":
                self._fill_row(ws, r, len(headers), GREEN_FILL)

        self._auto_width(ws)

    # ── Sheet 6: Vendor Summary ────────────────────────────────────────────────

    def _sheet_vendor_summary(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("06_Vendor_Summary")
        headers = ["Vendor ID", "Vendor Name", "Employee Count", "Total Regular Hours", "Total Overtime Hours", "Payroll Ready Count"]
        self._write_header_row(ws, headers)

        vendors = self.db.query(Vendor).all()
        for r, vendor in enumerate(vendors, 2):
            subs = (
                self.db.query(TimesheetSubmission)
                .filter(TimesheetSubmission.batch_id == batch_id, TimesheetSubmission.vendor_id == vendor.id)
                .all()
            )
            if not subs:
                continue
            entry_ids = [e.id for sub in subs for e in self.db.query(TimesheetEntry).filter(TimesheetEntry.submission_id == sub.id).all()]
            entries = self.db.query(TimesheetEntry).filter(TimesheetEntry.id.in_(entry_ids)).all() if entry_ids else []
            total_reg = sum(float(e.regular_hours or 0) for e in entries)
            total_ot = sum(float(e.overtime_hours or 0) for e in entries)
            ready = sum(1 for s in subs if s.payroll_status == "READY")

            ws.cell(r, 1, vendor.id)
            ws.cell(r, 2, vendor.name)
            ws.cell(r, 3, len(subs))
            ws.cell(r, 4, round(total_reg, 2))
            ws.cell(r, 5, round(total_ot, 2))
            ws.cell(r, 6, ready)

        self._auto_width(ws)

    # ── Sheet 7: Payroll Report ────────────────────────────────────────────────

    def _sheet_payroll_report(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("07_Payroll_Report")
        headers = [
            "Employee ID", "Employee Name", "Regular Hours", "OT Hours",
            "Regular Rate", "OT Rate", "Regular Pay", "OT Pay", "Total Pay",
            "Currency", "Payroll Status",
        ]
        self._write_header_row(ws, headers)

        submissions = (
            self.db.query(TimesheetSubmission)
            .filter(TimesheetSubmission.batch_id == batch_id)
            .all()
        )

        for r, sub in enumerate(submissions, 2):
            emp = self.db.query(Employee).filter(Employee.id == sub.employee_id).first()
            entries = self.db.query(TimesheetEntry).filter(TimesheetEntry.submission_id == sub.id).all()
            total_reg = sum(float(e.regular_hours or 0) for e in entries)
            total_ot = sum(float(e.overtime_hours or 0) for e in entries)

            # Get rate
            from app.db.models import EmployeeRate
            from datetime import date
            today = date.today()
            rate = (
                self.db.query(EmployeeRate)
                .filter(
                    EmployeeRate.employee_id == sub.employee_id,
                    EmployeeRate.effective_start_date <= today,
                )
                .order_by(EmployeeRate.effective_start_date.desc())
                .first()
            )

            reg_rate = float(rate.regular_rate) if rate else None
            ot_rate = float(rate.overtime_rate) if rate and rate.overtime_rate else (reg_rate * 1.5 if reg_rate else None)
            reg_pay = round(total_reg * reg_rate, 2) if reg_rate else None
            ot_pay = round(total_ot * ot_rate, 2) if ot_rate else None
            total_pay = round((reg_pay or 0) + (ot_pay or 0), 2) if reg_pay is not None else None

            ws.cell(r, 1, sub.employee_id or "")
            ws.cell(r, 2, emp.full_name if emp else "Unknown")
            ws.cell(r, 3, round(total_reg, 2))
            ws.cell(r, 4, round(total_ot, 2))
            ws.cell(r, 5, reg_rate or "MISSING")
            ws.cell(r, 6, ot_rate or "MISSING")
            ws.cell(r, 7, reg_pay or "MISSING")
            ws.cell(r, 8, ot_pay or "N/A")
            ws.cell(r, 9, total_pay or "MISSING")
            ws.cell(r, 10, "USD")
            ws.cell(r, 11, sub.payroll_status)

            if sub.payroll_status == "NOT_READY" or reg_rate is None:
                self._fill_row(ws, r, len(headers), RED_FILL)
            elif sub.payroll_status == "READY":
                self._fill_row(ws, r, len(headers), GREEN_FILL)

        self._auto_width(ws)

    # ── Sheet 8: ADP Export ────────────────────────────────────────────────────

    def _sheet_adp_export(self, wb: Workbook, batch_id: str) -> None:
        ws = wb.create_sheet("08_ADP_Export")
        # ADP-compatible columns
        headers = [
            "EMPLOYEE_ID", "EMPLOYEE_NAME", "PAY_PERIOD_START", "PAY_PERIOD_END",
            "REG_HOURS", "OT_HOURS", "REG_PAY", "OT_PAY", "TOTAL_PAY",
            "CURRENCY", "COST_CENTER", "DEPARTMENT",
        ]
        self._write_header_row(ws, headers)

        submissions = (
            self.db.query(TimesheetSubmission)
            .filter(
                TimesheetSubmission.batch_id == batch_id,
                TimesheetSubmission.payroll_status == "READY",
            )
            .all()
        )

        for r, sub in enumerate(submissions, 2):
            emp = self.db.query(Employee).filter(Employee.id == sub.employee_id).first()
            entries = self.db.query(TimesheetEntry).filter(TimesheetEntry.submission_id == sub.id).all()
            total_reg = sum(float(e.regular_hours or 0) for e in entries)
            total_ot = sum(float(e.overtime_hours or 0) for e in entries)

            from app.db.models import EmployeeRate
            from datetime import date
            rate = (
                self.db.query(EmployeeRate)
                .filter(
                    EmployeeRate.employee_id == sub.employee_id,
                    EmployeeRate.effective_start_date <= date.today(),
                )
                .order_by(EmployeeRate.effective_start_date.desc())
                .first()
            )
            reg_rate = float(rate.regular_rate) if rate else 0.0
            ot_rate = float(rate.overtime_rate) if rate and rate.overtime_rate else reg_rate * 1.5
            reg_pay = round(total_reg * reg_rate, 2)
            ot_pay = round(total_ot * ot_rate, 2)

            ws.cell(r, 1, emp.employee_code if emp and emp.employee_code else str(emp.id)[:8] if emp else "")
            ws.cell(r, 2, emp.full_name if emp else "")
            ws.cell(r, 3, str(sub.timesheet_start_date) if sub.timesheet_start_date else "")
            ws.cell(r, 4, str(sub.timesheet_end_date) if sub.timesheet_end_date else "")
            ws.cell(r, 5, round(total_reg, 2))
            ws.cell(r, 6, round(total_ot, 2))
            ws.cell(r, 7, reg_pay)
            ws.cell(r, 8, ot_pay)
            ws.cell(r, 9, round(reg_pay + ot_pay, 2))
            ws.cell(r, 10, "USD")
            ws.cell(r, 11, "")
            ws.cell(r, 12, "")

        self._auto_width(ws)

    # ── Helpers ────────────────────────────────────────────────────────────────

    def _write_header_row(self, ws, headers: list) -> None:
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _fill_row(ws, row: int, num_cols: int, fill) -> None:
        for col in range(1, num_cols + 1):
            ws.cell(row, col).fill = fill

    @staticmethod
    def _auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)
